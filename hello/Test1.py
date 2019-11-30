import tkinter as tk
from tkinter import ttk, messagebox, Text, filedialog
import sqlite3 as sql
from openpyxl import load_workbook as LoadBook
import threading


# base GUI Class
class GUI:
    def __init__(self, root, runCommand):
        mf = ttk.Frame(root, padding="5 5 5 5")
        mf.grid(column=0, row=0)
        mf.columnconfigure(0, weight=1)
        mf.rowconfigure(0, weight=1)

        # Global Values
        self.Fnm = tk.StringVar(root, "SearchFile.xlsx")
        self.Ncol = tk.StringVar(root, "D")
        self.Vcol = tk.StringVar(root, "C")
        self.PAcnt = tk.IntVar(root, 9)
        # Label
        tk.Label(mf, text="File Name").grid(column=1, row=1, pady=6)
        tk.Label(mf, text="Name Col").grid(column=1, row=2, pady=6)
        tk.Label(mf, text="Value Col").grid(column=3, row=2, pady=6)
        tk.Label(mf, text="Total PA").grid(column=5, row=2, pady=6)

        # components
        self.fname = ttk.Entry(mf, width=18, textvariable=self.Fnm)
        self.nmCol = ttk.Entry(mf, width=6, textvariable=self.Ncol)
        self.fsch = ttk.Button(mf, text="Browse", command=lambda: self.getFile(), width=7)
        self.valCol = ttk.Entry(mf, width=6, textvariable=self.Vcol)
        self.PAC = ttk.Entry(mf, width=6, textvariable=self.PAcnt)
        self.but = ttk.Button(mf, text="Refresh", command=runCommand)
        self.pgbar = ttk.Progressbar(mf, orient="horizontal", mode="determinate")
        self.log = Text(mf, width=40, height=10)

        # Design
        self.fname.grid(column=2, row=1, pady=3, columnspan=4, sticky='we')
        self.fsch.grid(column=6, row=1, pady=3, columnspan=1)
        self.nmCol.grid(column=2, row=2, pady=3)
        self.valCol.grid(column=4, row=2, pady=3)
        self.PAC.grid(column=6, row=2, pady=3)
        self.but.grid(column=3, row=3, columnspan=2, sticky='we')
        self.pgbar.grid(column=1, row=4, columnspan=6, sticky='we')
        self.log.grid(column=1, row=5, columnspan=6, pady=4)

    def refresh(self):
        pass

    def get(self):
        return [self.Fnm.get(), self.Ncol.get(), self.Vcol.get(), int(self.PAC.get())]

    def getFile(self):
        self.fname.delete(0, len(self.Fnm.get()))
        fnm = filedialog.askopenfilename(filetypes=(("Microsoft Excel", "*.xlsx"),
                                                    ("All files", "*.*")))
        try:
            self.fname.insert(0, fnm)
        except:  # <- naked except is a bad idea
           messagebox.showerror("Open Source File", "Failed to read file\n'%s'" % fnm)


# Base process Class
class Proc:
    def __init__(self, dets, pgbar, but, log):
        self.Fnm = dets[0]
        self.Ncol = dets[1]
        self.Vcol = dets[2]
        self.PAcnt = dets[3]
        self.pg = pgbar
        self.pg["maximum"] = self.PAcnt
        self.butt = but
        self.log = log

    def refresh(self):
        self.butt['state'] = 'disabled'
        self.log.insert('end', "Process Started...\n")
        self.conn = self.conDB('.\\DB\\medSort')
        self.cur = self.conn.cursor()
        try:
            dfile = LoadBook(self.Fnm)
            tblN = ['paSearchList', 'rcSearchList']
            ct = 0
            self.clearDB(tblN)
            for sheet in range(0, len(dfile.sheetnames)):
                self.log.insert('end', "\nWorking on datasheet: " + dfile.sheetnames[sheet])
                self.log.yview_pickplace("end")
                self.pg['value'] = (sheet)
                dfile._active_sheet_index = sheet
                dsheet = dfile.active
                if not sheet < self.PAcnt:
                    ct = 1
                nameSet, descSet = dsheet[self.Ncol][1:], dsheet[self.Vcol][1:]
                for nameData, valData in zip(nameSet, descSet):
                    if nameData.value is None:
                        break
                    self.insertIn([valData.value, nameData.value], tblN[ct])
            self.log.insert('end', "\nProcess Over...")
            self.log.yview_pickplace("end")
            messagebox.showinfo("Process Done",
                                "Search Database Refreshed!!\n Updated " + str(self.cur.lastrowid) + " rows")
        except IOError as e:
            messagebox.showwarning('Error', 'File not found!')
            self.log.insert('end', 'Please copy the required file to continue..')
        self.butt['state'] = 'enabled'

    def conDB(self, db):
        try:
            con = sql.connect(db, isolation_level=None)
            return con
        except sql.OperationalError:
            print('Operational Error,database not found')
            return None
            pass

    def clearDB(self, tbl):
        try:
            for _ in tbl:
                que = "Delete from " + tbl[0]
                self.cur.execute(que)
        except sql.OperationalError:
            pass

    def insertIn(self, values, tbl):
        try:
            que = "insert into " + tbl + " (contract, name) values (?,?)"
            self.cur.execute(que, values)
        except sql.OperationalError:
            que = "create table " + tbl + " (contract varchar(200), name varchar(200))"
            self.cur.execute(que)
            que = "insert into " + tbl + " (contract, name) values (?,?)"
            self.cur.execute(que, values)


# Base Application Class
class App:
    def __init__(self, master):
        self.master = master
        self.gui = GUI(self.master, self.runit)

    def runit(self):
        self.search = Proc(self.gui.get(), self.gui.pgbar, self.gui.but, self.gui.log)
        self.thread1 = threading.Thread(target=self.search.refresh)
        self.thread1.start()


def main():
    app = tk.Tk()
    gui = App(app)
    app.title("Refresh Search File")
    app.mainloop()


if __name__ == '__main__':
    main()